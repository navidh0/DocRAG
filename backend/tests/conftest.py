"""Pytest configuration and shared fixtures for RAG system tests."""
import pytest
from rest_framework.test import APIClient
from django.contrib.auth import get_user_model
from django.test import override_settings

User = get_user_model()


@pytest.fixture
def api_client_factory():
    """Factory fixture for creating multiple API clients."""
    def create_client(user=None):
        client = APIClient()
        if user:
            from rest_framework_simplejwt.tokens import RefreshToken
            refresh = RefreshToken.for_user(user)
            client.credentials(HTTP_AUTHORIZATION=f'Bearer {str(refresh.access_token)}')
            client.user = user
        return client
    return create_client


@pytest.fixture
def api_client():
    """API client for making requests to the REST endpoints."""
    return APIClient()


@pytest.fixture
def authenticated_client(test_user):  # ← inject test_user instead of creating a new one
    """API client with authenticated user."""
    client = APIClient()
    client.force_authenticate(user=test_user)
    client.user = test_user
    return client

@pytest.fixture
def test_user():
    """Create a test user."""
    return User.objects.create_user(
        username='testuser',
        email='test@example.com',
        password='TestPass123!'
    )


@pytest.fixture
def second_user():
    """Create a second test user for isolation testing."""
    return User.objects.create_user(
        username='otheruser',
        email='other@example.com',
        password='OtherPass123!'
    )


@pytest.fixture
def user_token(test_user):
    """Get JWT token for test user."""
    from rest_framework_simplejwt.tokens import RefreshToken
    refresh = RefreshToken.for_user(test_user)
    return str(refresh.access_token)


@pytest.fixture
def auth_headers(user_token):
    """Authorization headers with JWT token."""
    return {'HTTP_AUTHORIZATION': f'Bearer {user_token}'}


@pytest.fixture
def test_document_text():
    """Sample document text for testing."""
    return """Machine Learning Basics
Machine learning is a subset of artificial intelligence that enables systems to learn from data.

Natural Language Processing
NLP is a field of AI that focuses on processing and understanding human language.

Vector Embeddings
Embeddings represent text or data as numerical vectors in high-dimensional space.

Retrieval Augmented Generation
RAG combines retrieval systems with generative models for better answers.

Database Systems
PostgreSQL is a powerful open-source relational database with vector support."""


@pytest.fixture
def test_excel_bytes():
    """Sample Excel file bytes for testing."""
    import io
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Test Document'
    ws['A2'] = 'This is test data'
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@pytest.fixture
def test_pdf_bytes():
    """Sample PDF file bytes for testing."""
    try:
        from reportlab.pdfgen import canvas
        import io
        
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer)
        c.drawString(100, 750, "Test PDF Document")
        c.drawString(100, 700, "This is a test PDF for document upload testing")
        c.save()
        buffer.seek(0)
        return buffer.getvalue()
    except ImportError:
        # Fallback: return minimal PDF bytes if reportlab not available
        return b'%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj'


@pytest.fixture
def mock_ollama_client(monkeypatch):
    """Mock Ollama client for testing."""
    class MockOllamaClient:
        def embed(self, model, input):
            # Return fake embeddings
            if isinstance(input, list):
                return {'embeddings': [[0.1] * 768 for _ in input]}
            return {'embeddings': [[0.1] * 768]}
        
        def generate(self, model, prompt, stream=False):
            if stream:
                # Return generator for streaming
                def token_generator():
                    tokens = ["This", " is", " a", " test", " answer"]
                    for token in tokens:
                        yield {'response': token}
                return token_generator()
            return {'response': 'This is a test answer'}
    
    def mock_client(*args, **kwargs):
        return MockOllamaClient()
    
    monkeypatch.setattr('ollama.Client', mock_client)
    return MockOllamaClient()

@pytest.fixture(autouse=True)
def mock_embedding_task(mocker):
    """
    Patches Celery task dispatch globally across all tests.
    Prevents any test from accidentally hitting a real broker.
    Use `mock_embedding_task.assert_called_once_with(str(doc_id))`
    in upload tests to verify scheduling without executing.
    """
    return mocker.patch("documents.tasks.process_document_embedding.delay")


@pytest.fixture(autouse=True)
def reset_db():
    """Ensure database is clean between tests."""
    # This is handled by pytest-django's db fixture
    pass


# Django settings overrides for tests
@pytest.fixture(scope='session')
def django_db_setup(django_db_setup, django_db_blocker):
    """Customize database setup for tests."""
    # Use default SQLite for tests unless overridden
    pass

@pytest.fixture
def make_document(test_user):
    """
    Factory fixture for creating Document records directly in the DB,
    bypassing the upload API. Use this in tests that care about
    list/detail/delete/status behaviour, not upload behaviour.
    """
    def _make(file_name="test.pdf", file_type="pdf", status="pending", user=None, created_at=None):
        from documents.models import Document
        doc = Document.objects.create(
            user=user or test_user,
            file_name=file_name,
            file_type=file_type,
            status=status,
        )
        if created_at is not None:
            Document.objects.filter(id=doc.id).update(created_at=created_at)
            doc.refresh_from_db()
        return doc
    return _make